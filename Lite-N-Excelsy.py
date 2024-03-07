#
#   This script was written with Github Copilot
#
#   It scrapes the nutrition information from the Lite-N-Easy website and saves it to an Excel file
#   It was made for the purposes of ranking the meals by Sodium content for low-sodium diets.
#
#   DISCLAIMER:
#   It is the sole responsibility of you, the user to check the official Lite-N-Easy HTML pages
#   for accurate and up-to-date information regarding allergens and nutrition values.
#   The spreadsheet produced by this script has not been cross-referenced with the official
#   HTML pages for accuracy and is intended as a guide only for comparison purposes.
#
#   Any deaths resulting from the use of this script or the generated spreadsheet are not the
#   responsibility of the author and no correspondence will be entered into.
#
import requests
import openpyxl
import os
from lxml import html


#
#   CONFIGURATION
#

# set output file name
output_file = 'Lite-N-Easy.xlsx'

# set menu index URL
indexURL = 'https://www.liteneasy.com.au/ingredients-nutrition'

# set header row (change at own risk)
headers = ['Name', 'Serving size', 'Energy', 'Protein', 'Fat, Total',
           'Saturated Fat', 'Carbohydrate', 'Sugars', 'Fibre', 'Sodium', 'Ingredients']


#
#   FUNCTIONS
#

# function to fetch the HTML content from a URL and return a parsed HTML tree
def fetch_html_tree(url):
    response = requests.get(url)
    html_content = response.text
    tree = html.fromstring(html_content)
    return tree

#
#   MAIN
#


# fetch the HTML tree of index page
indexTree = fetch_html_tree(indexURL)

# initialise an array for the menu pages
pages = []

# get all links inside a table
indexMatches = indexTree.xpath('//table//a')
# for each link in the table
for a in indexMatches:
    # get the text from the link, flattening child elements
    linkTitle = ' '.join(a.itertext()).strip()
    # first clean the href attribute by stripping any parameters
    linkURL = a.attrib['href'].split('?')[0]
    # if the linkURL ends with .html and the title is not blank
    if linkURL.endswith('.html') and linkTitle:
        # append a dictionary with the title and URL to the pages array
        pages.append({'Title': linkTitle, 'URL': linkURL})


# create a new workbook
wb = openpyxl.Workbook()

# for each page in the array
for page in pages:
    # initialize a blank array to store the data
    data = []

    # create a new worksheet called the Title of the page
    ws = wb.create_sheet(title=page['Title'])

    # write the header row to the worksheet
    ws.append(headers)

    # edit the 'Sodium' header to append ' (mg)' as the units are stripped to allow sorting
    ws['J1'] = 'Sodium (mg)'

    # get HTML from current menu page
    currentURL = page['URL']
    currentTree = fetch_html_tree(currentURL)

    # Iterate over td elements with a h2 inside a span.IngredName
    # this corresponds to each meal on the page
    indexMatches = currentTree.xpath('//td[span[@class="IngredName"]//h2]')

    for td in indexMatches:
        # create a blank dictionary to store the data for this meal
        row = {}
        # find a h2 element inside the current table cell
        name = td.find('.//h2')
        # extract the text from the h2 element and set Name to this value
        row['Name'] = name.text if name is not None else ''
        # find a span.Ingred_Serving_Contents element inside the td
        # this is where the serving size lives
        serving_size = td.find('.//span[@class="Ingred_Serving_Contents"]')
        # extract the text from the span element and set Serving size to this value
        row['Serving size'] = serving_size.text if serving_size is not None else None
        # find all td elemts inside a table in the current td element
        # this is where the nutritional information lives
        nutris = td.xpath('.//table//td')
        # iterate in pairs of 3 (ignoring the per 100g column)
        for i in range(0, len(nutris), 3):
            # set the key to the text of the first td element
            key = ' '.join(nutris[i].itertext()).strip()
            # set the value to the text from the second td element
            value = ' '.join(nutris[i+1].itertext()).strip()
            # if the key is Sodium strip 'mg' from the value then strip any whitespace
            if key == 'Sodium':
                value = value.replace('mg', '').strip()
            # set the key and value in the data row dictionary
            row[key] = value
            # if the key was Sodium, exit the loop (end of table)
            if key == 'Sodium':
                break
        # extract the text from a span.Ingred_Ingred_Contents
        # this is where the ingredients live
        ings = td.find('.//span[@class="Ingred_Ingred_Contents"]')
        # flatten all child elements and just get the text
        ingsText = ' '.join(ings.itertext()).strip()
        row['Ingredients'] = ingsText
        # append the row dictionary to the data list
        data.append(row)

    # TODO: this could be improved to ensure that the data is written to the row
    # by comparing the keys to header rows. in its current form it relies on the
    # nutrition information being in the same order for every meal.

    # iterate over the data array
    for row in data:
        # convert the dictionary to a list
        rowlist = [row.get(key, '') for key in data[0]]
        # append the row to the worksheet
        ws.append(rowlist)


#
#   SAVE
#

# remove the default sheet
wb.remove(wb['Sheet'])


# If output file exists, delete it
if os.path.exists(output_file):
    os.remove(output_file)

# save the workbook to a file
wb.save(output_file)
